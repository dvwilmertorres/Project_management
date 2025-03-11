from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Pedir el título del proyecto al usuario
project_title = input("Ingrese el título del proyecto: ")

# Crear un archivo de Excel
wb = Workbook()
ws = wb.active
ws.title = "Project Plan"

# Escribir el título del proyecto en la celda A2
ws["A2"] = project_title

# Rellenar la información del Proyecto en las primeras filas
project_info = [
    ["PROJECT", "PROJECT DURATION", "PROJECT START DATE", "FINISH PROJECT DATE", "MODALITY"],
]

# Escribir la información del proyecto
for row in project_info:
    ws.append(row)

# Escribir las actividades y las fechas
stages_header = [
    "STAGES", "ACTIVITIES", "RESPONSIBLE", "DURATION", "START DATE", "FINISH DATE", "TASK STATUS",
    "11/20/2024", "11/21/2024", "11/22/2024", "11/25/2024", "11/26/2024", "11/27/2024", "11/28/2024",
    "11/29/2024", "12/2/2024", "12/3/2024", "12/4/2024", "12/5/2024", "12/6/2024", "12/9/2024",
    "12/10/2024", "12/11/2024", "12/12/2024", "12/13/2024", "12/16/2024", "12/17/2024", "12/18/2024",
    "12/19/2024", "12/20/2024", "12/23/2024", "12/24/2024", "12/25/2024", "12/26/2024", "12/27/2024",
    "12/30/2024", "12/31/2024", "1/1/2025", "1/2/2025", "1/3/2025", "1/6/2025", "1/7/2025", "1/8/2025",
    "1/9/2025", "1/10/2025", "1/13/2025", "1/14/2025"
]

# Escribir los encabezados de las actividades y fechas
ws.append(stages_header)

# Datos de las actividades
activities_data = [
    ["1-Start", "kick off", "Aaro", 1, "20/11/24", "20/11/24", "Completo"],
    ["2-Planning", "App Planning (Architecture)", "Paola Andrea Rodriguez Castro", 3, "20/11/24", "25/11/24", "En proceso"],
    ["2-Planning", "Cloud Design", "Paola Andrea Rodriguez Castro", 1, "25/11/24", "26/11/24", "En proceso"],
    ["2-Planning", "BackEnd Design", "Paola Andrea Rodriguez Castro", 3, "26/11/24", "29/11/24", "En proceso"],
    ["2-Planning", "Interface", "Paola Andrea Rodriguez Castro", 3, "29/11/24", "04/12/24", "En proceso"],
    ["2-Planning", "Deliverable: Architecture", "Paola Andrea Rodriguez Castro", 1, "04/12/24", "05/12/24", "En proceso"],
    ["2-Planning", "Project feasibility", "Aaro", 1, "05/12/24", "06/12/24", "Pendiente"],
    ["3-Execution", "Password Recovery Front", "Paola Andrea Rodriguez Castro", 2, "06/12/24", "10/12/24", "Pendiente"],
    ["3-Execution", "Password Recovery Back", "Juan", 3, "08/12/24", "11/12/24", "Pendiente"]
]

# Escribir las actividades
for activity in activities_data:
    ws.append(activity)

# Función para aplicar un color sólido a las celdas
def apply_solid_fill(start_col, end_col, row, start_color, end_color):
    solid_fill = PatternFill(start_color=start_color, end_color=end_color, fill_type="solid")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = solid_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

# Aplicar color sólido a los títulos de la información del proyecto
apply_solid_fill(1, 5, 1, "338aff", "92D050")  # Colores verdes suaves

# Aplicar color sólido a los títulos de las actividades y fechas
apply_solid_fill(1, len(stages_header), 2, "B6D7A8", "92D050")  # Colores verdes suaves

# Ajustar el ancho de las columnas
for col in range(1, len(stages_header) + 1):
    col_letter = get_column_letter(col)
    max_length = 0
    for row in ws.iter_rows(min_col=col, max_col=col):
        for cell in row:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[col_letter].width = adjusted_width

# Establecer el formato de las celdas
for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(stages_header)):
    for cell in row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Guardar el archivo de Excel
wb.save("project_plan_with_solid_fill.xlsx")
