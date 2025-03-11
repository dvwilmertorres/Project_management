import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill

def renombrar_proyecto(nombre_proyecto):
    # Agrega el sufijo '_gannt' al nombre del proyecto
    nuevo_nombre = nombre_proyecto + "_gannt"
    
    return nuevo_nombre

def crear_libro_excel(nuevo_nombre, nombre_proyecto):
    # Datos a insertar en la primera fila (project_info)
    project_info = [
        ["PROJECT", "PROJECT DURATION", "PROJECT START DATE", "FINISH PROJECT DATE", "MODALITY"],
    ]
    
    # Datos a insertar en la fila 5 (stages_header)
    stages_header = [
        ["STAGES", "ACTIVITIES", "RESPONSIBLE", "DURATION", "START DATE", "FINISH DATE", "TASK STATUS"]
    ]
    
    # Stages - colorimetry
    stages_list = [
        "Start", "Planning", "Execution", "Tests", 
        "Implementation/Deployment", "Stabilization/maintenance", "Project closure"
    ]
    
    # Convertir los datos en un DataFrame
    df_project_info = pd.DataFrame(project_info)
    df_stages_header = pd.DataFrame(stages_header)
    
    # Guardar el DataFrame como un archivo Excel
    with pd.ExcelWriter(f"{nuevo_nombre}.xlsx", engine="openpyxl") as writer:
        df_project_info.to_excel(writer, index=False, header=False, startrow=0, startcol=0)
        df_stages_header.to_excel(writer, index=False, header=False, startrow=4, startcol=0)

    # Cargar el archivo Excel recién creado
    wb = load_workbook(f"{nuevo_nombre}.xlsx")
    ws_roadmap = wb.active
    ws_roadmap.title = "ROADMAP"  # Renombrar la hoja 1 como "ROADMAP"

    # Crear la hoja "RESOURCES" y agregar la lista de Stages
    ws_resources = wb.create_sheet(title="RESOURCES")
    for i, stage in enumerate(stages_list, start=1):
        ws_resources[f"A{i}"] = stage
    
    # Colocar el nombre del proyecto en la celda A2 (sin el sufijo "_gannt")
    ws_roadmap["A2"] = nombre_proyecto
    
    # Definir los bordes (negros)
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Crear los estilos de fuente y fondo para los títulos
    font = Font(name="Aptos Narrow", bold=True, color="FFFFFF")
    fill = PatternFill(start_color="59a0ff", end_color="59a0ff", fill_type="solid")
    
    # Aplicar bordes y estilos a la fila de project_info (A1 a E1)
    for row in ws_roadmap.iter_rows(min_row=1, max_row=1, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            cell.font = font
            cell.fill = fill
    
    # Aplicar bordes y estilos a la fila de stages_header (A5 a G5)
    for row in ws_roadmap.iter_rows(min_row=5, max_row=5, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.font = font
            cell.fill = fill
    
    # Ahora hacer que las celdas de la columna A de "ROADMAP" (A6 a A500) hagan referencia a la lista de etapas de "RESOURCES"
    for i in range(6, 501):  # Desde A6 hasta A500
        # Asignar la referencia de la celda en la hoja "RESOURCES"
        ws_roadmap[f"A{i}"] = f"=RESOURCES!A{((i-6) % len(stages_list)) + 1}"
    
    # Guardar el archivo con los bordes, fuente, fondo y el nombre en A2
